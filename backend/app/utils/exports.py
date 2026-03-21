"""
Export Utilities for Reports
ParcelFlow - Multi-tenant Logistics Platform

Provides Excel (XLSX) and PDF generation for various reports.
"""
import io
from datetime import date, datetime
from typing import List, Dict, Any, Optional
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from weasyprint import HTML, CSS
from jinja2 import Environment, FileSystemLoader
import os


# ==================== EXCEL EXPORT UTILITIES ====================

class ExcelExporter:
    """Excel export utility class using openpyxl"""
    
    # Color definitions
    HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TOTAL_FILL = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    TOTAL_FONT = Font(bold=True, size=11)
    BORDER = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    def __init__(self, title: str, business_name: str = "ParcelFlow"):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.title = title
        self.business_name = business_name
        self.current_row = 1
    
    def add_title(self, title: str = None, subtitle: str = None):
        """Add report title and optional subtitle"""
        title = title or self.title
        
        # Title row
        self.worksheet.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.worksheet.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(bold=True, size=16, color="1F2937")
        cell.alignment = Alignment(horizontal='center')
        self.current_row += 1
        
        # Subtitle (date range or description)
        if subtitle:
            self.worksheet.merge_cells(f'A{self.current_row}:H{self.current_row}')
            cell = self.worksheet.cell(row=self.current_row, column=1, value=subtitle)
            cell.font = Font(size=11, color="6B7280")
            cell.alignment = Alignment(horizontal='center')
            self.current_row += 1
        
        # Business name
        self.worksheet.merge_cells(f'A{self.current_row}:H{self.current_row}')
        cell = self.worksheet.cell(row=self.current_row, column=1, value=self.business_name)
        cell.font = Font(size=10, color="9CA3AF")
        cell.alignment = Alignment(horizontal='center')
        self.current_row += 2  # Empty row after header
    
    def add_headers(self, headers: List[str], start_col: int = 1):
        """Add table headers with styling"""
        for col, header in enumerate(headers, start=start_col):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=header)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.BORDER
        
        self.current_row += 1
        return self.current_row - 1  # Return header row for later reference
    
    def add_row(self, data: List[Any], start_col: int = 1, is_total: bool = False):
        """Add a data row with optional total styling"""
        for col, value in enumerate(data, start=start_col):
            cell = self.worksheet.cell(row=self.current_row, column=col, value=value)
            cell.border = self.BORDER
            
            if is_total:
                cell.fill = self.TOTAL_FILL
                cell.font = self.TOTAL_FONT
            
            # Format numbers
            if isinstance(value, (int, float, Decimal)):
                cell.alignment = Alignment(horizontal='right')
                if isinstance(value, float):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = Alignment(horizontal='left')
        
        self.current_row += 1
    
    def add_summary_section(self, title: str, data: Dict[str, Any]):
        """Add a summary section with key-value pairs"""
        # Section title
        cell = self.worksheet.cell(row=self.current_row, column=1, value=title)
        cell.font = Font(bold=True, size=12, color="1F2937")
        self.current_row += 1
        
        # Key-value pairs
        for key, value in data.items():
            cell_key = self.worksheet.cell(row=self.current_row, column=1, value=key)
            cell_key.font = Font(color="6B7280")
            
            cell_value = self.worksheet.cell(row=self.current_row, column=2, value=value)
            cell_value.font = Font(bold=True)
            if isinstance(value, (int, float, Decimal)):
                cell_value.number_format = '#,##0.00'
            
            self.current_row += 1
        
        self.current_row += 1  # Empty row
    
    def auto_adjust_columns(self, min_width: int = 12, max_width: int = 50):
        """Auto-adjust column widths based on content"""
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            self.worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def add_chart(self, chart_type: str, data_range: str, title: str, position: str = "E2"):
        """Add a chart to the worksheet"""
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "line":
            chart = LineChart()
        else:
            return
        
        chart.title = title
        # Chart configuration would depend on data structure
        # This is a placeholder for more complex charting needs
    
    def get_bytes(self) -> bytes:
        """Get the workbook as bytes for download"""
        self.auto_adjust_columns()
        
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()


# ==================== PDF EXPORT UTILITIES ====================

class PDFExporter:
    """PDF export utility class using WeasyPrint"""
    
    def __init__(self, template_dir: str = None):
        if template_dir is None:
            # Default to backend templates
            template_dir = os.path.join(os.path.dirname(__file__), '..', 'templates', 'pdf')
        
        self.template_dir = template_dir
        self.env = None
        
        if os.path.exists(template_dir):
            self.env = Environment(loader=FileSystemLoader(template_dir))
    
    def generate_from_html(self, html_content: str, styles: str = None) -> bytes:
        """Generate PDF from HTML content"""
        default_styles = """
        <style>
            @page {
                size: A4 landscape;
                margin: 1.5cm;
                @top-right {
                    content: "ParcelFlow Report";
                    font-size: 9px;
                    color: #9CA3AF;
                }
                @bottom-center {
                    content: "Page " counter(page) " of " counter(pages);
                    font-size: 9px;
                    color: #9CA3AF;
                }
            }
            
            body {
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                font-size: 11px;
                color: #1F2937;
                line-height: 1.4;
            }
            
            .report-header {
                text-align: center;
                margin-bottom: 20px;
                border-bottom: 2px solid #1F2937;
                padding-bottom: 15px;
            }
            
            .report-title {
                font-size: 20px;
                font-weight: bold;
                color: #1F2937;
                margin: 0;
            }
            
            .report-subtitle {
                font-size: 12px;
                color: #6B7280;
                margin: 5px 0 0 0;
            }
            
            .business-name {
                font-size: 10px;
                color: #9CA3AF;
                margin-top: 5px;
            }
            
            .summary-grid {
                display: grid;
                grid-template-columns: repeat(4, 1fr);
                gap: 15px;
                margin-bottom: 20px;
            }
            
            .summary-card {
                border: 1px solid #E5E7EB;
                border-radius: 8px;
                padding: 12px;
                text-align: center;
            }
            
            .summary-label {
                font-size: 10px;
                color: #6B7280;
                text-transform: uppercase;
            }
            
            .summary-value {
                font-size: 18px;
                font-weight: bold;
                margin-top: 5px;
            }
            
            .summary-value.green { color: #16A34A; }
            .summary-value.red { color: #DC2626; }
            .summary-value.blue { color: #2563EB; }
            .summary-value.yellow { color: #D97706; }
            
            table {
                width: 100%;
                border-collapse: collapse;
                margin-top: 15px;
            }
            
            th {
                background-color: #1F2937;
                color: white;
                padding: 10px 8px;
                text-align: left;
                font-size: 10px;
                text-transform: uppercase;
            }
            
            td {
                padding: 8px;
                border-bottom: 1px solid #E5E7EB;
            }
            
            tr:nth-child(even) {
                background-color: #F9FAFB;
            }
            
            tr.total-row {
                background-color: #E5E7EB;
                font-weight: bold;
            }
            
            .text-right { text-align: right; }
            .text-center { text-align: center; }
            
            .footer {
                margin-top: 30px;
                text-align: center;
                font-size: 9px;
                color: #9CA3AF;
                border-top: 1px solid #E5E7EB;
                padding-top: 10px;
            }
        </style>
        """
        
        full_html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            {styles or default_styles}
        </head>
        <body>
            {html_content}
        </body>
        </html>
        """
        
        pdf_bytes = HTML(string=full_html).write_pdf()
        return pdf_bytes
    
    def generate_sales_report(self, report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
        """Generate a formatted sales report PDF"""
        html_content = f"""
        <div class="report-header">
            <h1 class="report-title">Sales Report</h1>
            <p class="report-subtitle">{report_data.get('start_date', '')} to {report_data.get('end_date', '')}</p>
            <p class="business-name">{business_name}</p>
        </div>
        
        <div class="summary-grid">
            <div class="summary-card">
                <p class="summary-label">Total Orders</p>
                <p class="summary-value blue">{report_data.get('total_orders', 0):,}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Total Revenue</p>
                <p class="summary-value green">${report_data.get('total_revenue', 0):,.2f}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Delivery Fees</p>
                <p class="summary-value">${report_data.get('total_delivery_fees', 0):,.2f}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">COD Collected</p>
                <p class="summary-value yellow">${report_data.get('total_cod', 0):,.2f}</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th class="text-right">Orders</th>
                    <th class="text-right">Revenue</th>
                    <th class="text-right">Delivery Fees</th>
                    <th class="text-right">COD Collected</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"""
                <tr>
                    <td>{item.get('date', '')}</td>
                    <td class="text-right">{item.get('orders', 0):,}</td>
                    <td class="text-right">${item.get('revenue', 0):,.2f}</td>
                    <td class="text-right">${item.get('delivery_fees', 0):,.2f}</td>
                    <td class="text-right">${item.get('cod_collected', 0):,.2f}</td>
                </tr>
                """ for item in report_data.get('items', [])])}
                <tr class="total-row">
                    <td>TOTALS</td>
                    <td class="text-right">{report_data.get('total_orders', 0):,}</td>
                    <td class="text-right">${report_data.get('total_revenue', 0):,.2f}</td>
                    <td class="text-right">${report_data.get('total_delivery_fees', 0):,.2f}</td>
                    <td class="text-right">${report_data.get('total_cod', 0):,.2f}</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ParcelFlow Logistics Platform</p>
        </div>
        """
        
        return self.generate_from_html(html_content)
    
    def generate_delivery_report(self, report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
        """Generate a formatted delivery performance report PDF"""
        html_content = f"""
        <div class="report-header">
            <h1 class="report-title">Delivery Performance Report</h1>
            <p class="report-subtitle">{report_data.get('start_date', '')} to {report_data.get('end_date', '')}</p>
            <p class="business-name">{business_name}</p>
        </div>
        
        <div class="summary-grid">
            <div class="summary-card">
                <p class="summary-label">Total Deliveries</p>
                <p class="summary-value blue">{report_data.get('total_deliveries', 0):,}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Successful</p>
                <p class="summary-value green">{report_data.get('total_successful', 0):,}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Failed</p>
                <p class="summary-value red">{report_data.get('total_failed', 0):,}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Success Rate</p>
                <p class="summary-value yellow">{report_data.get('overall_success_rate', 0):.1f}%</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th class="text-right">Total</th>
                    <th class="text-right">Delivered</th>
                    <th class="text-right">Failed</th>
                    <th class="text-right">Returned</th>
                    <th class="text-right">Success Rate</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"""
                <tr>
                    <td>{item.get('date', '')}</td>
                    <td class="text-right">{item.get('total', 0):,}</td>
                    <td class="text-right" style="color: #16A34A;">{item.get('delivered', 0):,}</td>
                    <td class="text-right" style="color: #DC2626;">{item.get('failed', 0):,}</td>
                    <td class="text-right" style="color: #D97706;">{item.get('returned', 0):,}</td>
                    <td class="text-right">{item.get('success_rate', 0):.1f}%</td>
                </tr>
                """ for item in report_data.get('items', [])])}
                <tr class="total-row">
                    <td>TOTALS</td>
                    <td class="text-right">{report_data.get('total_deliveries', 0):,}</td>
                    <td class="text-right">{report_data.get('total_successful', 0):,}</td>
                    <td class="text-right">{report_data.get('total_failed', 0):,}</td>
                    <td class="text-right">-</td>
                    <td class="text-right">{report_data.get('overall_success_rate', 0):.1f}%</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ParcelFlow Logistics Platform</p>
        </div>
        """
        
        return self.generate_from_html(html_content)
    
    def generate_agent_report(self, report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
        """Generate a formatted agent performance report PDF"""
        html_content = f"""
        <div class="report-header">
            <h1 class="report-title">Agent Performance Report</h1>
            <p class="report-subtitle">{report_data.get('period_start', '')} to {report_data.get('period_end', '')}</p>
            <p class="business-name">{business_name}</p>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Agent Name</th>
                    <th class="text-right">Total</th>
                    <th class="text-right">Successful</th>
                    <th class="text-right">Failed</th>
                    <th class="text-right">Success Rate</th>
                    <th class="text-right">Rating</th>
                    <th class="text-right">COD Collected</th>
                    <th class="text-right">Commissions</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"""
                <tr>
                    <td>{item.get('agent_name', '')}</td>
                    <td class="text-right">{item.get('total_deliveries', 0):,}</td>
                    <td class="text-right" style="color: #16A34A;">{item.get('successful', 0):,}</td>
                    <td class="text-right" style="color: #DC2626;">{item.get('failed', 0):,}</td>
                    <td class="text-right">{item.get('success_rate', 0):.1f}%</td>
                    <td class="text-right">{item.get('rating', 0):.1f}</td>
                    <td class="text-right">${item.get('cod_collected', 0):,.2f}</td>
                    <td class="text-right">${item.get('commissions_earned', 0):,.2f}</td>
                </tr>
                """ for item in report_data.get('items', [])])}
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ParcelFlow Logistics Platform</p>
        </div>
        """
        
        return self.generate_from_html(html_content)
    
    def generate_vendor_report(self, report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
        """Generate a formatted vendor settlement report PDF"""
        html_content = f"""
        <div class="report-header">
            <h1 class="report-title">Vendor Settlement Report</h1>
            <p class="report-subtitle">{report_data.get('start_date', '')} to {report_data.get('end_date', '')}</p>
            <p class="business-name">{business_name}</p>
        </div>
        
        <div class="summary-grid">
            <div class="summary-card">
                <p class="summary-label">Total Commissions</p>
                <p class="summary-value blue">${report_data.get('total_commissions', 0):,.2f}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Total Due</p>
                <p class="summary-value yellow">${report_data.get('total_due', 0):,.2f}</p>
            </div>
            <div class="summary-card">
                <p class="summary-label">Total Paid</p>
                <p class="summary-value green">${report_data.get('total_paid', 0):,.2f}</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Vendor Name</th>
                    <th class="text-right">Orders</th>
                    <th class="text-right">Total Value</th>
                    <th class="text-right">Commission %</th>
                    <th class="text-right">Commission</th>
                    <th class="text-right">Amount Due</th>
                    <th class="text-right">Paid</th>
                    <th class="text-right">Balance</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"""
                <tr>
                    <td>{item.get('vendor_name', '')}</td>
                    <td class="text-right">{item.get('total_orders', 0):,}</td>
                    <td class="text-right">${item.get('total_value', 0):,.2f}</td>
                    <td class="text-right">{item.get('commission_rate', 0):.1f}%</td>
                    <td class="text-right">${item.get('commission_amount', 0):,.2f}</td>
                    <td class="text-right">${item.get('amount_due', 0):,.2f}</td>
                    <td class="text-right">${item.get('amount_paid', 0):,.2f}</td>
                    <td class="text-right" style="color: {'#DC2626' if item.get('balance', 0) > 0 else '#16A34A'};">
                        ${abs(item.get('balance', 0)):,.2f}
                    </td>
                </tr>
                """ for item in report_data.get('items', [])])}
                <tr class="total-row">
                    <td>TOTALS</td>
                    <td class="text-right">-</td>
                    <td class="text-right">-</td>
                    <td class="text-right">-</td>
                    <td class="text-right">${report_data.get('total_commissions', 0):,.2f}</td>
                    <td class="text-right">${report_data.get('total_due', 0):,.2f}</td>
                    <td class="text-right">${report_data.get('total_paid', 0):,.2f}</td>
                    <td class="text-right">-</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ParcelFlow Logistics Platform</p>
        </div>
        """
        
        return self.generate_from_html(html_content)
    
    def generate_expense_report(self, report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
        """Generate a formatted expense report PDF"""
        html_content = f"""
        <div class="report-header">
            <h1 class="report-title">Expense Summary Report</h1>
            <p class="report-subtitle">{report_data.get('period_start', '')} to {report_data.get('period_end', '')}</p>
            <p class="business-name">{business_name}</p>
        </div>
        
        <div class="summary-grid">
            <div class="summary-card" style="grid-column: span 4;">
                <p class="summary-label">Total Expenses</p>
                <p class="summary-value red">${report_data.get('total_expenses', 0):,.2f}</p>
            </div>
        </div>
        
        <table>
            <thead>
                <tr>
                    <th>Category</th>
                    <th class="text-right">Total Amount</th>
                    <th class="text-right">Count</th>
                    <th class="text-right">Percentage</th>
                </tr>
            </thead>
            <tbody>
                {''.join([f"""
                <tr>
                    <td>{item.get('category', '')}</td>
                    <td class="text-right" style="color: #DC2626;">${item.get('total_amount', 0):,.2f}</td>
                    <td class="text-right">{item.get('count', 0):,}</td>
                    <td class="text-right">{item.get('percentage', 0):.1f}%</td>
                </tr>
                """ for item in report_data.get('items', [])])}
                <tr class="total-row">
                    <td>TOTAL</td>
                    <td class="text-right">${report_data.get('total_expenses', 0):,.2f}</td>
                    <td class="text-right">-</td>
                    <td class="text-right">100%</td>
                </tr>
            </tbody>
        </table>
        
        <div class="footer">
            <p>Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | ParcelFlow Logistics Platform</p>
        </div>
        """
        
        return self.generate_from_html(html_content)


# ==================== HELPER FUNCTIONS ====================

def create_sales_excel(report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
    """Create Excel file for sales report"""
    exporter = ExcelExporter("Sales Report", business_name)
    
    # Add title
    exporter.add_title(
        "Sales Report",
        f"{report_data.get('start_date', '')} to {report_data.get('end_date', '')}"
    )
    
    # Add summary section
    exporter.add_summary_section("Summary", {
        "Total Orders": report_data.get('total_orders', 0),
        "Total Revenue": f"${report_data.get('total_revenue', 0):,.2f}",
        "Delivery Fees": f"${report_data.get('total_delivery_fees', 0):,.2f}",
        "COD Collected": f"${report_data.get('total_cod', 0):,.2f}",
        "Avg Order Value": f"${report_data.get('average_order_value', 0):,.2f}"
    })
    
    # Add table headers
    exporter.add_headers(["Date", "Orders", "Revenue", "Delivery Fees", "COD Collected"])
    
    # Add data rows
    for item in report_data.get('items', []):
        exporter.add_row([
            item.get('date', ''),
            item.get('orders', 0),
            item.get('revenue', 0),
            item.get('delivery_fees', 0),
            item.get('cod_collected', 0)
        ])
    
    # Add totals row
    exporter.add_row([
        "TOTALS",
        report_data.get('total_orders', 0),
        report_data.get('total_revenue', 0),
        report_data.get('total_delivery_fees', 0),
        report_data.get('total_cod', 0)
    ], is_total=True)
    
    return exporter.get_bytes()


def create_delivery_excel(report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
    """Create Excel file for delivery performance report"""
    exporter = ExcelExporter("Delivery Performance Report", business_name)
    
    exporter.add_title(
        "Delivery Performance Report",
        f"{report_data.get('start_date', '')} to {report_data.get('end_date', '')}"
    )
    
    exporter.add_summary_section("Summary", {
        "Total Deliveries": report_data.get('total_deliveries', 0),
        "Successful": report_data.get('total_successful', 0),
        "Failed": report_data.get('total_failed', 0),
        "Success Rate": f"{report_data.get('overall_success_rate', 0):.1f}%",
        "Avg Delivery Time": f"{report_data.get('average_delivery_time_hours', 0):.1f} hours" if report_data.get('average_delivery_time_hours') else "N/A"
    })
    
    exporter.add_headers(["Date", "Total", "Delivered", "Failed", "Returned", "Success Rate"])
    
    for item in report_data.get('items', []):
        exporter.add_row([
            item.get('date', ''),
            item.get('total', 0),
            item.get('delivered', 0),
            item.get('failed', 0),
            item.get('returned', 0),
            f"{item.get('success_rate', 0):.1f}%"
        ])
    
    exporter.add_row([
        "TOTALS",
        report_data.get('total_deliveries', 0),
        report_data.get('total_successful', 0),
        report_data.get('total_failed', 0),
        "-",
        f"{report_data.get('overall_success_rate', 0):.1f}%"
    ], is_total=True)
    
    return exporter.get_bytes()


def create_agent_excel(report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
    """Create Excel file for agent performance report"""
    exporter = ExcelExporter("Agent Performance Report", business_name)
    
    exporter.add_title(
        "Agent Performance Report",
        f"{report_data.get('period_start', '')} to {report_data.get('period_end', '')}"
    )
    
    exporter.add_headers([
        "Agent Name", "Total", "Successful", "Failed", "Success Rate", 
        "Rating", "COD Collected", "Commissions"
    ])
    
    for item in report_data.get('items', []):
        exporter.add_row([
            item.get('agent_name', ''),
            item.get('total_deliveries', 0),
            item.get('successful', 0),
            item.get('failed', 0),
            f"{item.get('success_rate', 0):.1f}%",
            item.get('rating', 0),
            item.get('cod_collected', 0),
            item.get('commissions_earned', 0)
        ])
    
    return exporter.get_bytes()


def create_vendor_excel(report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
    """Create Excel file for vendor settlement report"""
    exporter = ExcelExporter("Vendor Settlement Report", business_name)
    
    exporter.add_title(
        "Vendor Settlement Report",
        f"{report_data.get('start_date', '')} to {report_data.get('end_date', '')}"
    )
    
    exporter.add_summary_section("Summary", {
        "Total Commissions": f"${report_data.get('total_commissions', 0):,.2f}",
        "Total Due": f"${report_data.get('total_due', 0):,.2f}",
        "Total Paid": f"${report_data.get('total_paid', 0):,.2f}"
    })
    
    exporter.add_headers([
        "Vendor Name", "Orders", "Total Value", "Commission %",
        "Commission", "Amount Due", "Paid", "Balance"
    ])
    
    for item in report_data.get('items', []):
        exporter.add_row([
            item.get('vendor_name', ''),
            item.get('total_orders', 0),
            item.get('total_value', 0),
            f"{item.get('commission_rate', 0):.1f}%",
            item.get('commission_amount', 0),
            item.get('amount_due', 0),
            item.get('amount_paid', 0),
            item.get('balance', 0)
        ])
    
    exporter.add_row([
        "TOTALS", "-", "-", "-",
        report_data.get('total_commissions', 0),
        report_data.get('total_due', 0),
        report_data.get('total_paid', 0),
        "-"
    ], is_total=True)
    
    return exporter.get_bytes()


def create_expense_excel(report_data: Dict[str, Any], business_name: str = "ParcelFlow") -> bytes:
    """Create Excel file for expense report"""
    exporter = ExcelExporter("Expense Summary Report", business_name)
    
    exporter.add_title(
        "Expense Summary Report",
        f"{report_data.get('period_start', '')} to {report_data.get('period_end', '')}"
    )
    
    exporter.add_summary_section("Summary", {
        "Total Expenses": f"${report_data.get('total_expenses', 0):,.2f}"
    })
    
    exporter.add_headers(["Category", "Total Amount", "Count", "Percentage"])
    
    for item in report_data.get('items', []):
        exporter.add_row([
            item.get('category', ''),
            item.get('total_amount', 0),
            item.get('count', 0),
            f"{item.get('percentage', 0):.1f}%"
        ])
    
    exporter.add_row([
        "TOTAL",
        report_data.get('total_expenses', 0),
        "-",
        "100%"
    ], is_total=True)
    
    return exporter.get_bytes()
